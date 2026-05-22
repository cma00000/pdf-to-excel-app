#!/usr/bin/env python3
"""PDF Table → Excel  —  desktop app"""

import os
import platform
import subprocess
import threading
import tkinter as tk
from tkinter import ttk, filedialog

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    _DND = True
except ImportError:
    _DND = False

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── Palette ───────────────────────────────────────────────────────────────────
BG      = "#f5f5f7"
HEADER  = "#1F4E79"
ACCENT  = "#2E86C1"
SUCCESS = "#1e8449"
ERROR   = "#c0392b"
ZONE_BD = "#b0c4d8"

# ── PDF extraction ────────────────────────────────────────────────────────────
def extract_tables(pdf_path):
    tables  = []
    pending = None

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for raw in (page.extract_tables() or []):
                if not raw or all(
                    all(c is None or str(c).strip() == "" for c in row) for row in raw
                ):
                    continue

                clean = [[str(c or "").strip() for c in row] for row in raw]

                if pending is not None:
                    if pending[0] == clean[0]:      # same headers → multi-page continuation
                        pending.extend(clean[1:])
                        continue
                    tables.append(pending)
                    pending = None

                pending = clean

        if pending:
            tables.append(pending)

    return tables


def write_excel(tables, out_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")
    alt_fill = PatternFill("solid", fgColor="D6E4F0")

    for i, rows in enumerate(tables):
        ws = wb.create_sheet(f"Table {i + 1}")

        for r_idx, row in enumerate(rows):
            for c_idx, val in enumerate(row):
                cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=val or "")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if r_idx == 0:
                    cell.font  = hdr_font
                    cell.fill  = hdr_fill
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                elif r_idx % 2 == 0:
                    cell.fill = alt_fill

        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(w + 4, 50)

        ws.freeze_panes = "A2"

    wb.save(out_path)


# ── UI ────────────────────────────────────────────────────────────────────────
class App:
    def __init__(self):
        Root = TkinterDnD.Tk if _DND else tk.Tk
        self.root = Root()
        self.root.title("PDF Table → Excel")
        self.root.resizable(False, False)
        self.root.configure(bg=BG)
        self._busy = False
        self._build()
        self._center(440, 340)

    def _center(self, w, h):
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        self.root.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")

    def _build(self):
        # Header strip
        hdr = tk.Frame(self.root, bg=HEADER, height=54)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="PDF Table → Excel", bg=HEADER, fg="white",
                 font=("Helvetica Neue", 15, "bold")).pack(side="left", padx=18, pady=14)
        tk.Label(hdr, text="tables only · no prose", bg=HEADER, fg="#7aaecf",
                 font=("Helvetica Neue", 11)).pack(side="right", padx=18)

        # Main area
        body = tk.Frame(self.root, bg=BG)
        body.pack(fill="both", expand=True, padx=20, pady=18)

        # Drop zone
        self.zone = tk.Frame(body, bg="white",
                              highlightthickness=2,
                              highlightbackground=ZONE_BD,
                              highlightcolor=ACCENT)
        self.zone.pack(fill="both", expand=True)

        pivot = tk.Frame(self.zone, bg="white")
        pivot.place(relx=.5, rely=.45, anchor="center")

        tk.Label(pivot, text="📄", bg="white", font=("", 38)).pack()
        self.drop_label = tk.Label(pivot, text="Drop a PDF here",
                                    bg="white", fg="#222",
                                    font=("Helvetica Neue", 14))
        self.drop_label.pack(pady=(8, 2))
        tk.Label(pivot, text="or", bg="white", fg="#aaa",
                 font=("Helvetica Neue", 11)).pack()
        _hand_cursor = "hand2" if platform.system() != "Darwin" else "pointinghand"
        browse = tk.Label(pivot, text="click to browse", bg="white",
                          fg=ACCENT, font=("Helvetica Neue", 11, "underline"),
                          cursor=_hand_cursor)
        browse.pack()

        for w in (self.zone, self.drop_label, browse):
            w.bind("<Button-1>", lambda _: self._open_dialog())

        if _DND:
            self.zone.drop_target_register(DND_FILES)
            self.zone.dnd_bind("<<Drop>>", self._on_drop)

        # Progress bar (hidden until needed)
        self.bar = ttk.Progressbar(body, mode="indeterminate", length=400)

        # Status label
        self.status_var = tk.StringVar()
        self.status_lbl = tk.Label(body, textvariable=self.status_var,
                                    bg=BG, fg="#555",
                                    font=("Helvetica Neue", 11),
                                    wraplength=400, justify="left")
        self.status_lbl.pack(fill="x", pady=(10, 0))

    # ── Interactions ──────────────────────────────────────────────────────────
    def _open_dialog(self):
        if self._busy:
            return
        path = filedialog.askopenfilename(
            title="Select a PDF", filetypes=[("PDF files", "*.pdf")]
        )
        if path:
            self._process(path)

    def _on_drop(self, event):
        if self._busy:
            return
        path = event.data.strip().strip("{}")   # tkinterdnd2 wraps spaced paths in {}
        if path.lower().endswith(".pdf"):
            self._process(path)
        else:
            self._set_status("Please drop a PDF file.", ERROR)

    # ── Processing ────────────────────────────────────────────────────────────
    def _process(self, pdf_path):
        self._busy = True
        self._set_status(f"Extracting tables from {os.path.basename(pdf_path)}…", "#555")
        self.bar.pack(fill="x", pady=(8, 0))
        self.bar.start(10)

        def run():
            try:
                tables = extract_tables(pdf_path)
                if not tables:
                    self.root.after(0, lambda: self._finish(None, 0, "No tables found in this PDF."))
                    return
                out = os.path.splitext(pdf_path)[0] + ".xlsx"
                write_excel(tables, out)
                self.root.after(0, lambda: self._finish(out, len(tables), None))
            except Exception as e:
                self.root.after(0, lambda: self._finish(None, 0, str(e)))

        threading.Thread(target=run, daemon=True).start()

    def _finish(self, out_path, n, error):
        self.bar.stop()
        self.bar.pack_forget()
        self._busy = False

        if error:
            self._set_status(f"✗  {error}", ERROR)
        else:
            name = os.path.basename(out_path)
            self._set_status(f"✓  {n} table(s) → {name}", SUCCESS)
            if platform.system() == "Windows":
                os.startfile(out_path)
            elif platform.system() == "Darwin":
                subprocess.Popen(["open", out_path])
            else:
                subprocess.Popen(["xdg-open", out_path])

    def _set_status(self, msg, color="#555"):
        self.status_var.set(msg)
        self.status_lbl.config(fg=color)

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    App().run()
